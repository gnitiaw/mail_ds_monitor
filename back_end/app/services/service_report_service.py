from __future__ import annotations

import csv
import html
import re
from datetime import datetime, timezone
from io import BytesIO, StringIO
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.exceptions import ConflictError, ParamError
from app.models.enums import (
    CompletenessStatus,
    PeriodRule,
    ReportRunStatus,
    ReportType,
    SectionDataStatus,
    SourceRunStatus,
    SourceType,
)
from app.models.service_report import ServiceReportConfig, ServiceReportRun, ServiceReportSourceRun
from app.models.user import User
from app.schemas.common import ErrorCode

SOURCE_SPECS: dict[str, dict[str, Any]] = {
    SourceType.INSPECTION.value: {
        "required": ("inspection_date", "system_name", "status"),
        "aliases": {
            "inspection_date": ("inspection_date", "date", "巡检日期", "日期"),
            "system_name": ("system_name", "system", "host", "系统名称", "巡检对象"),
            "status": ("status", "result", "状态", "巡检结果"),
            "issue_count": ("issue_count", "issues", "异常数", "问题数"),
            "summary": ("summary", "remark", "description", "摘要", "说明"),
        },
    },
    SourceType.VULNERABILITY.value: {
        "required": ("vulnerability_id", "severity", "status"),
        "aliases": {
            "vulnerability_id": ("vulnerability_id", "vuln_id", "漏洞编号", "漏洞id"),
            "severity": ("severity", "level", "risk_level", "严重级别", "风险等级"),
            "status": ("status", "fix_status", "修复状态", "状态"),
            "fixed_at": ("fixed_at", "closed_at", "修复时间", "关闭时间"),
            "system_name": ("system_name", "system", "host", "系统名称"),
            "summary": ("summary", "remark", "description", "摘要", "说明"),
        },
    },
    SourceType.WORKLOG.value: {
        "required": ("work_date", "category", "summary"),
        "aliases": {
            "work_date": ("work_date", "date", "工作日期", "日期"),
            "category": ("category", "type", "工作类别", "类别"),
            "hours": ("hours", "duration", "工时", "耗时"),
            "summary": ("summary", "content", "工作内容", "摘要"),
            "result": ("result", "outcome", "处理结果", "结果"),
        },
    },
    SourceType.ZENTAO_BUG.value: {
        "required": ("bug_id", "severity", "status", "title"),
        "aliases": {
            "bug_id": ("bug_id", "id", "缺陷编号", "bug编号"),
            "severity": ("severity", "level", "严重程度", "级别"),
            "status": ("status", "bug_status", "状态"),
            "title": ("title", "bug_title", "标题"),
            "closed_at": ("closed_at", "resolved_at", "关闭时间", "解决时间"),
        },
    },
}

REPORT_TEMPLATES = {
    ReportType.MONTHLY.value: "ops_service_monthly_v1",
    ReportType.QUARTERLY.value: "ops_service_quarterly_v1",
    ReportType.ANNUAL.value: "ops_service_annual_v1",
}

SECTIONS = {
    ReportType.MONTHLY.value: [
        ("executive_summary", "执行摘要", None),
        ("inspection_overview", "巡检概览", SourceType.INSPECTION.value),
        ("vulnerability_fix", "漏洞修复情况", SourceType.VULNERABILITY.value),
        ("worklog_summary", "运维工作记录", SourceType.WORKLOG.value),
        ("zentao_defects", "禅道缺陷处理", SourceType.ZENTAO_BUG.value),
        ("risk_and_next_steps", "风险与下阶段计划", None),
    ],
    ReportType.QUARTERLY.value: [
        ("executive_summary", "季度执行摘要", None),
        ("quarterly_service_overview", "季度服务概览", None),
        ("quarterly_inspection_summary", "季度巡检总结", SourceType.INSPECTION.value),
        ("quarterly_vulnerability_trend", "季度漏洞治理趋势", SourceType.VULNERABILITY.value),
        ("quarterly_worklog_highlights", "季度运维重点工作", SourceType.WORKLOG.value),
        ("quarterly_defect_trend", "季度缺陷处理趋势", SourceType.ZENTAO_BUG.value),
        ("next_quarter_focus", "下季度重点计划", None),
    ],
    ReportType.ANNUAL.value: [
        ("executive_summary", "年度执行摘要", None),
        ("annual_service_overview", "年度服务概览", None),
        ("annual_inspection_summary", "年度巡检总结", SourceType.INSPECTION.value),
        ("annual_vulnerability_governance", "年度漏洞治理", SourceType.VULNERABILITY.value),
        ("annual_worklog_highlights", "年度重点运维工作", SourceType.WORKLOG.value),
        ("annual_defect_summary", "年度缺陷处理总结", SourceType.ZENTAO_BUG.value),
        ("next_year_plan", "下一年度计划", None),
    ],
}


class ServiceReportService:
    @staticmethod
    def validate_config_payload(db: Session, payload: Any) -> None:
        for user_id in (
            payload.project_owner_user_id,
            payload.template_owner_user_id,
            payload.metric_owner_user_id,
        ):
            user = db.get(User, user_id)
            if not user or not user.is_active:
                raise ParamError("owner user does not exist or is inactive")

        if REPORT_TEMPLATES[payload.report_type] != payload.template_key:
            raise ParamError("template_key does not match report_type")

        expected_rule = {
            ReportType.MONTHLY.value: PeriodRule.NATURAL_MONTH.value,
            ReportType.QUARTERLY.value: PeriodRule.NATURAL_QUARTER.value,
            ReportType.ANNUAL.value: PeriodRule.NATURAL_YEAR.value,
        }[payload.report_type]
        if payload.period_rule not in (expected_rule, PeriodRule.CUSTOM.value):
            raise ParamError("period_rule does not match report_type")

    @staticmethod
    def enforce_single_project(db: Session, project_name: str) -> None:
        active_configs = list(
            db.scalars(select(ServiceReportConfig).where(ServiceReportConfig.enabled == True)).all()  # noqa: E712
        )
        if active_configs and any(item.project_name != project_name for item in active_configs):
            raise ConflictError(
                "V1 仅允许单项目试点，当前已有其他启用中的项目配置",
                code=ErrorCode.SINGLE_PROJECT_CONFLICT,
            )

    @staticmethod
    def create_config(db: Session, payload: Any) -> ServiceReportConfig:
        ServiceReportService.validate_config_payload(db, payload)
        if payload.enabled:
            ServiceReportService.enforce_single_project(db, payload.project_name)
        existing = db.scalar(select(ServiceReportConfig).where(ServiceReportConfig.name == payload.name))
        if existing:
            raise ConflictError("report config name already exists")

        config = ServiceReportConfig(
            name=payload.name,
            project_name=payload.project_name,
            report_type=payload.report_type,
            period_rule=payload.period_rule,
            template_key=payload.template_key,
            project_owner_user_id=payload.project_owner_user_id,
            template_owner_user_id=payload.template_owner_user_id,
            metric_owner_user_id=payload.metric_owner_user_id,
            enabled=payload.enabled,
            recipient_emails=payload.recipient_emails,
            source_bindings=payload.source_bindings,
        )
        db.add(config)
        db.commit()
        db.refresh(config)
        return config

    @staticmethod
    def list_configs(
        db: Session,
        report_type: str | None,
        enabled: bool | None,
        keyword: str | None,
        page: int,
        page_size: int,
    ) -> tuple[list[ServiceReportConfig], int]:
        items = list(db.scalars(select(ServiceReportConfig).order_by(ServiceReportConfig.created_at.desc())).all())
        if report_type:
            items = [item for item in items if item.report_type == report_type]
        if enabled is not None:
            items = [item for item in items if item.enabled is enabled]
        if keyword:
            token = keyword.lower()
            items = [item for item in items if token in item.name.lower() or token in item.project_name.lower()]
        total = len(items)
        start = (page - 1) * page_size
        return items[start:start + page_size], total

    @staticmethod
    async def create_source_run(
        db: Session,
        config: ServiceReportConfig,
        window_start: datetime,
        window_end: datetime,
        files: dict[str, UploadFile | None],
    ) -> ServiceReportSourceRun:
        source_run = ServiceReportSourceRun(
            config_id=config.id,
            window_start=window_start,
            window_end=window_end,
            status=SourceRunStatus.RUNNING.value,
            included_sources=list(SOURCE_SPECS.keys()),
            source_results=[],
            snapshot_payload={"sources": {}},
        )
        db.add(source_run)
        db.commit()
        db.refresh(source_run)

        summaries: list[dict[str, Any]] = []
        payload_sources: dict[str, Any] = {}
        any_success = False
        any_partial = False
        for source_type in SOURCE_SPECS:
            result = await ServiceReportService._parse_upload(source_type, files.get(source_type))
            summaries.append(result["summary"])
            payload_sources[source_type] = result["payload"]
            if result["summary"]["status"] == SourceRunStatus.SUCCESS.value:
                any_success = True
            if result["summary"]["status"] == SourceRunStatus.PARTIAL_SUCCESS.value:
                any_success = True
                any_partial = True

        if any(item["status"] == SourceRunStatus.FAILED.value for item in summaries):
            source_run.status = SourceRunStatus.PARTIAL_SUCCESS.value if any_success else SourceRunStatus.FAILED.value
        else:
            source_run.status = SourceRunStatus.PARTIAL_SUCCESS.value if any_partial else SourceRunStatus.SUCCESS.value

        source_run.source_results = summaries
        source_run.snapshot_payload = {"sources": payload_sources}
        source_run.finished_at = datetime.now(timezone.utc)
        db.commit()
        db.refresh(source_run)
        return source_run

    @staticmethod
    def create_report_run(
        db: Session,
        config: ServiceReportConfig,
        source_run: ServiceReportSourceRun,
        window_start: datetime,
        window_end: datetime,
        force_regenerate: bool,
    ) -> ServiceReportRun:
        if source_run.config_id != config.id:
            raise ParamError("source run does not belong to this config")
        if not force_regenerate:
            existing = db.scalar(
                select(ServiceReportRun)
                .where(ServiceReportRun.config_id == config.id)
                .where(ServiceReportRun.source_run_id == source_run.id)
                .where(ServiceReportRun.status.in_([ReportRunStatus.RUNNING.value, ReportRunStatus.SUCCESS.value]))
            )
            if existing:
                return existing

        summary = ServiceReportService._build_source_summary(source_run)
        payload, completeness, evidence_refs = ServiceReportService._build_report(config.report_type, config.project_name, summary)
        run = ServiceReportRun(
            config_id=config.id,
            source_run_id=source_run.id,
            window_start=window_start,
            window_end=window_end,
            status=ReportRunStatus.SUCCESS.value,
            completeness_status=completeness,
            config_snapshot=ServiceReportService.config_snapshot(config),
            source_snapshot_summary=summary,
            report_payload=payload,
            manual_note=None,
            export_artifacts=[],
            evidence_refs=evidence_refs,
            error_message=None,
            finished_at=datetime.now(timezone.utc),
        )
        db.add(run)
        db.commit()
        db.refresh(run)
        return run

    @staticmethod
    def list_runs(db: Session, config_id: str | None, report_type: str | None, status: str | None, page: int, page_size: int) -> tuple[list[dict[str, Any]], int]:
        runs = list(db.scalars(select(ServiceReportRun).order_by(ServiceReportRun.created_at.desc())).all())
        configs = {item.id: item for item in db.scalars(select(ServiceReportConfig)).all()}
        items: list[dict[str, Any]] = []
        for run in runs:
            config = configs.get(run.config_id)
            if not config:
                continue
            if config_id and run.config_id != config_id:
                continue
            if report_type and config.report_type != report_type:
                continue
            if status and run.status != status:
                continue
            items.append(
                {
                    "run_id": run.id,
                    "config_id": run.config_id,
                    "config_name": config.name,
                    "project_name": config.project_name,
                    "report_type": config.report_type,
                    "status": run.status,
                    "completeness_status": run.completeness_status,
                    "window_start": run.window_start,
                    "window_end": run.window_end,
                    "source_run_id": run.source_run_id,
                    "export_formats": [artifact["format"] for artifact in (run.export_artifacts or [])],
                    "created_at": run.created_at,
                    "finished_at": run.finished_at,
                }
            )
        total = len(items)
        start = (page - 1) * page_size
        return items[start:start + page_size], total

    @staticmethod
    def update_manual_note(db: Session, run: ServiceReportRun, manual_note: str | None) -> ServiceReportRun:
        run.manual_note = manual_note
        db.commit()
        db.refresh(run)
        return run

    @staticmethod
    def register_export(db: Session, run: ServiceReportRun, export_format: str, overwrite: bool) -> dict[str, Any]:
        if run.status != ReportRunStatus.SUCCESS.value:
            raise ConflictError("current report is not ready to export")
        if run.completeness_status == CompletenessStatus.BLOCKED.value:
            raise ConflictError("当前报告 completeness_status=blocked，禁止导出", code=ErrorCode.EXPORT_BLOCKED)
        artifacts = list(run.export_artifacts or [])
        existing = next((item for item in artifacts if item["format"] == export_format), None)
        if existing and not overwrite:
            return existing
        artifact = {
            "format": export_format,
            "file_name": ServiceReportService.export_filename(run, export_format),
            "download_url": f"/api/v1/service-report-runs/{run.id}/export?format={export_format}",
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }
        artifacts = [item for item in artifacts if item["format"] != export_format]
        artifacts.append(artifact)
        run.export_artifacts = artifacts
        db.commit()
        db.refresh(run)
        return artifact

    @staticmethod
    def render_export(run: ServiceReportRun, export_format: str) -> str:
        payload = run.report_payload or {}
        lines: list[str] = []
        if run.completeness_status == CompletenessStatus.PARTIAL.value:
            lines.extend(["仅供内部复核", ""])
        lines.extend([payload.get("summary_markdown", ""), ""])
        for section in payload.get("sections", []):
            lines.extend([f"## {section['title']}", ""])
            if section["data_status"] == SectionDataStatus.PARTIAL.value:
                lines.extend(["> 章节数据不完整，仅供内部复核", ""])
            if section["data_status"] == SectionDataStatus.BLOCKED.value and section.get("blocking_reason"):
                lines.extend([f"> 阻塞原因：{section['blocking_reason']}", ""])
            lines.extend([section["content_markdown"], ""])
        if run.manual_note:
            lines.extend(["## 人工补充说明", "", run.manual_note, ""])
        markdown = "\n".join(lines)
        if export_format == "markdown":
            return markdown
        return "<html><body><pre style='white-space: pre-wrap; font-family: sans-serif;'>" + html.escape(markdown) + "</pre></body></html>"

    @staticmethod
    def config_snapshot(config: ServiceReportConfig) -> dict[str, Any]:
        return {
            "name": config.name,
            "project_name": config.project_name,
            "report_type": config.report_type,
            "period_rule": config.period_rule,
            "template_key": config.template_key,
            "project_owner_user_id": config.project_owner_user_id,
            "template_owner_user_id": config.template_owner_user_id,
            "metric_owner_user_id": config.metric_owner_user_id,
            "recipient_emails": config.recipient_emails,
            "source_bindings": config.source_bindings,
        }

    @staticmethod
    async def _parse_upload(source_type: str, upload: UploadFile | None) -> dict[str, Any]:
        if upload is None:
            return ServiceReportService._failed_source(source_type, "missing file")
        rows = await ServiceReportService._load_rows(upload)
        spec = SOURCE_SPECS[source_type]
        normalized: list[dict[str, Any]] = []
        errors: list[str] = []
        for idx, row in enumerate(rows, start=2):
            item = ServiceReportService._normalize(row, spec["aliases"])
            missing = [field for field in spec["required"] if not item.get(field)]
            if missing:
                errors.append(f"row {idx}: missing {', '.join(missing)}")
                continue
            normalized.append(ServiceReportService._post_process(source_type, item))
        if rows and not normalized and errors:
            status = SourceRunStatus.FAILED.value
        elif errors:
            status = SourceRunStatus.PARTIAL_SUCCESS.value
        else:
            status = SourceRunStatus.SUCCESS.value
        return {
            "summary": {
                "source_type": source_type,
                "status": status,
                "record_count": len(rows),
                "valid_row_count": len(normalized),
                "invalid_row_count": len(errors),
                "error_message": errors[0] if errors else None,
            },
            "payload": {
                "records": normalized,
                "errors": errors[:20],
                "evidence_refs": ServiceReportService._evidence_refs(source_type, normalized),
            },
        }

    @staticmethod
    async def _load_rows(upload: UploadFile) -> list[dict[str, str]]:
        content = await upload.read()
        name = (upload.filename or "").lower()
        if name.endswith(".csv"):
            return [dict(row) for row in csv.DictReader(StringIO(content.decode("utf-8-sig")))]
        if name.endswith(".xlsx"):
            workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            raw_rows = list(sheet.iter_rows(values_only=True))
            if not raw_rows:
                return []
            headers = [str(cell).strip() if cell is not None else "" for cell in raw_rows[0]]
            items: list[dict[str, str]] = []
            for row in raw_rows[1:]:
                item = {}
                for index, header in enumerate(headers):
                    if header:
                        value = row[index] if index < len(row) else None
                        item[header] = "" if value is None else str(value).strip()
                items.append(item)
            return items
        raise ParamError("only .csv and .xlsx files are supported")

    @staticmethod
    def _normalize(row: dict[str, Any], aliases: dict[str, tuple[str, ...]]) -> dict[str, Any]:
        lower_row = {str(key).strip().lower(): value for key, value in row.items() if key is not None}
        result: dict[str, Any] = {}
        for field, candidates in aliases.items():
            value = None
            for candidate in candidates:
                value = lower_row.get(candidate.lower())
                if value not in (None, ""):
                    break
            result[field] = value if value not in (None, "") else None
        return result

    @staticmethod
    def _post_process(source_type: str, item: dict[str, Any]) -> dict[str, Any]:
        data = dict(item)
        if source_type == SourceType.INSPECTION.value:
            data["issue_count"] = ServiceReportService._to_int(data.get("issue_count"))
        elif source_type == SourceType.VULNERABILITY.value:
            data["fixed"] = str(data.get("status", "")).lower() in {"fixed", "closed", "已修复", "已关闭"}
        elif source_type == SourceType.WORKLOG.value:
            data["hours"] = ServiceReportService._to_float(data.get("hours"))
        elif source_type == SourceType.ZENTAO_BUG.value:
            data["closed"] = str(data.get("status", "")).lower() in {"closed", "resolved", "已关闭", "已解决"}
        return data

    @staticmethod
    def _evidence_refs(source_type: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        refs: list[dict[str, Any]] = []
        for row in rows[:5]:
            ref_id = row.get("vulnerability_id") or row.get("bug_id") or row.get("inspection_date") or row.get("work_date") or "ref"
            refs.append(
                {
                    "source_type": source_type,
                    "ref_type": "record",
                    "ref_id": str(ref_id),
                    "title": row.get("title") or row.get("summary") or row.get("system_name") or row.get("category"),
                    "summary": row.get("summary") or row.get("status"),
                }
            )
        return refs

    @staticmethod
    def _build_source_summary(source_run: ServiceReportSourceRun) -> dict[str, Any]:
        sources = (source_run.snapshot_payload or {}).get("sources", {})
        inspection = sources.get(SourceType.INSPECTION.value, {}).get("records", [])
        vulnerabilities = sources.get(SourceType.VULNERABILITY.value, {}).get("records", [])
        worklogs = sources.get(SourceType.WORKLOG.value, {}).get("records", [])
        bugs = sources.get(SourceType.ZENTAO_BUG.value, {}).get("records", [])
        evidence_refs: list[dict[str, Any]] = []
        for source_type in SOURCE_SPECS:
            evidence_refs.extend(sources.get(source_type, {}).get("evidence_refs", []))
        return {
            "overview": {
                "inspection_total": len(inspection),
                "inspection_exception_total": sum(ServiceReportService._to_int(row.get("issue_count")) for row in inspection),
                "vulnerability_total": len(vulnerabilities),
                "vulnerability_fixed_total": sum(1 for row in vulnerabilities if row.get("fixed")),
                "vulnerability_unfixed_total": sum(1 for row in vulnerabilities if not row.get("fixed")),
                "worklog_total": len(worklogs),
                "zentao_bug_total": len(bugs),
                "zentao_bug_closed_total": sum(1 for row in bugs if row.get("closed")),
                "high_risk_count": sum(1 for row in vulnerabilities + bugs if str(row.get("severity", "")).lower() in {"high", "critical", "高", "严重"}),
            },
            "source_results": source_run.source_results or [],
            "evidence_refs": evidence_refs[:12],
        }

    @staticmethod
    def _build_report(report_type: str, project_name: str, source_summary: dict[str, Any]) -> tuple[dict[str, Any], str, list[dict[str, Any]]]:
        source_status = {item["source_type"]: item for item in source_summary.get("source_results", [])}
        sections: list[dict[str, Any]] = []
        statuses: list[str] = []
        overview = source_summary.get("overview", {})
        for key, title, source_type in SECTIONS[report_type]:
            data_status, reason = ServiceReportService._section_status(source_type, source_status)
            statuses.append(data_status)
            sections.append(
                {
                    "key": key,
                    "title": title,
                    "data_status": data_status,
                    "blocking_reason": reason,
                    "content_markdown": ServiceReportService._section_content(key, title, data_status, project_name, overview, source_status),
                }
            )
        completeness = CompletenessStatus.READY.value
        if SectionDataStatus.BLOCKED.value in statuses:
            completeness = CompletenessStatus.BLOCKED.value
        elif SectionDataStatus.PARTIAL.value in statuses:
            completeness = CompletenessStatus.PARTIAL.value
        summary = {
            ReportType.MONTHLY.value: "月报",
            ReportType.QUARTERLY.value: "季报",
            ReportType.ANNUAL.value: "年报",
        }[report_type]
        payload = {
            "summary_markdown": (
                f"# {project_name}{summary}\n\n"
                f"- 巡检记录数：{overview.get('inspection_total', 0)}\n"
                f"- 漏洞总数：{overview.get('vulnerability_total', 0)}，已修复 {overview.get('vulnerability_fixed_total', 0)}\n"
                f"- 运维工作记录：{overview.get('worklog_total', 0)}\n"
                f"- 禅道缺陷总数：{overview.get('zentao_bug_total', 0)}，已关闭 {overview.get('zentao_bug_closed_total', 0)}\n"
                f"- 报告完整度：{completeness}\n"
            ),
            "sections": sections,
        }
        return payload, completeness, source_summary.get("evidence_refs", [])

    @staticmethod
    def _section_status(source_type: str | None, source_status: dict[str, dict[str, Any]]) -> tuple[str, str | None]:
        if source_type is None:
            if any(item.get("status") == SourceRunStatus.FAILED.value for item in source_status.values()):
                return SectionDataStatus.BLOCKED.value, "存在必需数据源解析失败"
            if any(item.get("status") == SourceRunStatus.PARTIAL_SUCCESS.value for item in source_status.values()):
                return SectionDataStatus.PARTIAL.value, "存在部分数据降级"
            return SectionDataStatus.READY.value, None
        item = source_status.get(source_type)
        if not item or item.get("status") == SourceRunStatus.FAILED.value:
            return SectionDataStatus.BLOCKED.value, f"{source_type} 数据不可用"
        if item.get("status") == SourceRunStatus.PARTIAL_SUCCESS.value:
            return SectionDataStatus.PARTIAL.value, f"{source_type} 数据部分降级"
        return SectionDataStatus.READY.value, None

    @staticmethod
    def _section_content(key: str, title: str, data_status: str, project_name: str, overview: dict[str, Any], source_status: dict[str, dict[str, Any]]) -> str:
        if data_status == SectionDataStatus.BLOCKED.value:
            return f"{title} 暂无法生成，请补齐对应数据源后重试。"
        if key in {"executive_summary", "quarterly_service_overview", "annual_service_overview"}:
            return f"{project_name} 当前窗口内完成巡检 {overview.get('inspection_total', 0)} 次，处理漏洞 {overview.get('vulnerability_total', 0)} 项，记录运维工作 {overview.get('worklog_total', 0)} 条，跟踪缺陷 {overview.get('zentao_bug_total', 0)} 条。"
        if "inspection" in key:
            return f"巡检记录 {overview.get('inspection_total', 0)} 条，累计发现问题 {overview.get('inspection_exception_total', 0)} 项。"
        if "vulnerability" in key:
            return f"漏洞总数 {overview.get('vulnerability_total', 0)}，已修复 {overview.get('vulnerability_fixed_total', 0)}，未修复 {overview.get('vulnerability_unfixed_total', 0)}。"
        if "worklog" in key:
            return f"运维工作记录 {overview.get('worklog_total', 0)} 条，建议结合人工补充说明突出重点事项。"
        if "defect" in key or "zentao" in key:
            return f"禅道缺陷总数 {overview.get('zentao_bug_total', 0)}，已关闭 {overview.get('zentao_bug_closed_total', 0)}。"
        if key in {"risk_and_next_steps", "next_quarter_focus", "next_year_plan"}:
            has_partial = any(item.get("status") == SourceRunStatus.PARTIAL_SUCCESS.value for item in source_status.values())
            return f"当前高风险项 {overview.get('high_risk_count', 0)} 项。{'存在部分数据降级，建议内部复核后再外发。' if has_partial else '建议持续跟踪高风险漏洞、关键缺陷和重点事项。'}"
        return f"{title} 已生成。"

    @staticmethod
    def export_filename(run: ServiceReportRun, export_format: str) -> str:
        snapshot = run.config_snapshot or {}
        suffix = "md" if export_format == "markdown" else "html"
        project_name = ServiceReportService._safe_filename_part(snapshot.get("project_name"))
        report_type = ServiceReportService._safe_filename_part(snapshot.get("report_type"))
        return f"{project_name}-{report_type}-{run.window_start.date().isoformat()}.{suffix}"

    @staticmethod
    def _failed_source(source_type: str, message: str) -> dict[str, Any]:
        return {
            "summary": {
                "source_type": source_type,
                "status": SourceRunStatus.FAILED.value,
                "record_count": 0,
                "valid_row_count": 0,
                "invalid_row_count": 0,
                "error_message": message,
            },
            "payload": {"records": [], "errors": [message], "evidence_refs": []},
        }

    @staticmethod
    def _to_int(value: Any) -> int:
        try:
            return int(float(value)) if value not in (None, "") else 0
        except (TypeError, ValueError):
            return 0

    @staticmethod
    def _to_float(value: Any) -> float:
        try:
            return float(value) if value not in (None, "") else 0.0
        except (TypeError, ValueError):
            return 0.0

    @staticmethod
    def _safe_filename_part(value: Any) -> str:
        text = str(value or "").strip().lower()
        text = re.sub(r"[^a-z0-9._-]+", "-", text)
        text = text.strip("-")
        return text or "service-report"
